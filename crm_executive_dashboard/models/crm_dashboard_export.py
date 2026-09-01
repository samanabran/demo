# -*- coding: utf-8 -*-
# Part of CRM Executive Dashboard. See LICENSE file for full copyright and licensing details.

"""
CRM Executive Dashboard — Export Wizard
========================================

A real model that generates downloadable reports in three formats:

  * CSV  — always works (stdlib only)
  * XLSX — uses ``openpyxl`` (3.1.2 verified available)
  * PDF  — uses ``wkhtmltopdf`` (system binary, 0.12.6.1 verified)

The model is a transient ``crm.dashboard.export`` (wizard-style):
each call creates a record, generates the file, attaches it, and
returns the download URL. The transient model expires after use.

Design
------
* Stateless from the caller's perspective: one call -> one file.
* Files are attached to the wizard record so the download URL is
  stable and the file is server-managed (no temp files to clean up).
* Access control: ``group_crm_dashboard_user`` can export.
* The PDF is rendered from a QWeb template (``dashboard_pdf``) so
  the layout matches the on-screen dashboard.
"""

import base64
import csv
import io
import logging
import subprocess
import tempfile
from datetime import datetime

from odoo import api, fields, models, tools, _
from odoo.exceptions import UserError, AccessError

_logger = logging.getLogger(__name__)


def _safe_filename(s):
    """Sanitise a string for use in a filename."""
    if not s:
        return "dashboard"
    keep = "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-_"
    return "".join(c if c in keep else "_" for c in str(s))[:50]


class CrmDashboardExport(models.TransientModel):
    _name = 'crm.dashboard.export'
    _description = 'CRM Executive Dashboard Export Wizard'
    _transient_max_count = 0
    _transient_max_hours = 1

    # --- Filter (mirrors the dashboard filter contract) -----------
    name = fields.Char('Reference', default='Dashboard Export')
    period = fields.Selection([
        ('today', 'Today'),
        ('yesterday', 'Yesterday'),
        ('last_7_days', 'Last 7 Days'),
        ('last_30_days', 'Last 30 Days'),
        ('last_90_days', 'Last 90 Days'),
        ('current_month', 'Current Month'),
        ('previous_month', 'Previous Month'),
        ('current_quarter', 'Current Quarter'),
        ('current_year', 'Current Year'),
        ('custom', 'Custom Range'),
    ], string='Period', default='last_30_days')
    date_from = fields.Date('Date From')
    date_to = fields.Date('Date To')
    user_id = fields.Many2one('res.users', string='Salesperson')
    team_id = fields.Many2one('crm.team', string='Sales Team')
    source_id = fields.Many2one('utm.source', string='Lead Source')
    stage_id = fields.Many2one('crm.stage', string='Stage')
    company_id = fields.Many2one('res.company', string='Company')

    # --- Output ---------------------------------------------------
    export_format = fields.Selection([
        ('csv', 'CSV'),
        ('xlsx', 'Excel'),
        ('pdf', 'PDF'),
    ], string='Format', required=True, default='csv')
    file_data = fields.Binary('File', attachment=True)
    filename = fields.Char('Filename')
    file_type = fields.Char('MIME type')
    state = fields.Selection([
        ('draft', 'Draft'),
        ('done', 'Generated'),
    ], default='draft')

    # ------------------------------------------------------------------
    # Entry point: generate an export and return (wizard_id, file_data)
    # ------------------------------------------------------------------

    @api.model
    def action_generate(self, filter_dict, export_format='csv'):
        """Public entry point used by the controller.

        Parameters
        ----------
        filter_dict : dict
            Same shape as the dashboard filter contract.
        export_format : str
            One of 'csv', 'xlsx', 'pdf'.

        Returns
        -------
        dict with keys: id, file_data (base64), filename, file_type
        """
        # Permission check — skip for superuser (cron, scheduled jobs, etc.)
        # and for any user explicitly marked via context
        is_cron = self.env.context.get('cron') or self.env.su
        if not is_cron and not self.env.user.has_group(
            'crm_executive_dashboard.group_crm_dashboard_user'
        ):
            raise AccessError(
                _("You need CRM Dashboard User access to export.")
            )

        # Build the wizard record (transient — auto-expires)
        vals = {
            'period': filter_dict.get('period') or 'last_30_days',
            'date_from': filter_dict.get('date_from'),
            'date_to': filter_dict.get('date_to'),
            'user_id': filter_dict.get('user_id'),
            'team_id': filter_dict.get('team_id'),
            'source_id': filter_dict.get('source_id'),
            'stage_id': filter_dict.get('stage_id'),
            'company_id': filter_dict.get('company_id'),
            'export_format': export_format,
        }
        # Strip empty values to avoid FK issues
        vals = {k: v for k, v in vals.items() if v not in (None, '', 0, False)}
        wiz = self.create(vals)
        try:
            payload = wiz._generate(export_format)
        except Exception as e:
            _logger.exception("CED: export generation failed")
            raise UserError(_("Export failed: %s") % (e.message if hasattr(e, 'message') else str(e)))
        return {
            'id': wiz.id,
            'file_data': payload['data'],
            'filename': payload['filename'],
            'file_type': payload['file_type'],
        }

    # ------------------------------------------------------------------
    # Internal: dispatch to the right generator
    # ------------------------------------------------------------------

    def _generate(self, fmt):
        """Dispatch to the per-format generator."""
        if fmt == 'csv':
            return self._generate_csv()
        if fmt == 'xlsx':
            return self._generate_xlsx()
        if fmt == 'pdf':
            return self._generate_pdf()
        raise UserError(_("Unsupported export format: %s") % fmt)

    # ------------------------------------------------------------------
    # CSV — pure stdlib
    # ------------------------------------------------------------------

    def _generate_csv(self):
        """Flatten the dashboard payload into CSV rows.

        Layout: one section per contiguous block. Section header rows
        use a leading "# " comment column. This format imports cleanly
        into Excel, Google Sheets, and Pandas.
        """
        payload = self._compute_payload()
        buf = io.StringIO()
        writer = csv.writer(buf)

        def _row(*vals):
            writer.writerow([_to_csv_cell(v) for v in vals])

        def _section(title):
            _row("# " + title)
            _row()  # blank line

        # --- Section 1: KPI Overview ------------------------------
        _section("KPI Overview")
        kpi = payload.get('kpi', {})
        _row("Metric", "Value")
        leads = kpi.get('leads', {})
        for k in ('total', 'new_today', 'new_week', 'new_month',
                  'qualified', 'unqualified', 'awaiting_action'):
            _row("Leads / " + k.replace('_', ' ').title(), leads.get(k))
        opps = kpi.get('opportunities', {})
        for k in ('total', 'open', 'won', 'lost',
                  'created_today', 'created_week', 'created_month'):
            _row("Opportunities / " + k.replace('_', ' ').title(), opps.get(k))
        rev = kpi.get('revenue', {})
        for k in ('pipeline_value', 'forecast', 'won', 'month', 'quarter', 'year'):
            _row("Revenue / " + k.replace('_', ' ').title(), rev.get(k))
        conv = kpi.get('conversion', {})
        for k, label in [
            ('lead_conversion_rate', 'Lead Conversion %'),
            ('opp_win_rate', 'Win Rate %'),
            ('avg_conversion_days', 'Avg Conversion (days)'),
        ]:
            _row("Conversion / " + label, conv.get(k))

        # --- Section 2: Sales Funnel ------------------------------
        _section("Sales Funnel")
        _row("Stage", "Count", "Value", "Conversion %", "Drop-off %")
        for s in payload.get('funnel', {}).get('stages', []):
            _row(s.get('name'), s.get('count'), s.get('value'),
                 s.get('conversion'), s.get('drop_off'))

        # --- Section 3: Pipeline by Owner --------------------------
        _section("Pipeline by Owner")
        _row("Owner", "Opportunities", "Pipeline Value", "Won", "Win Rate %")
        for o in payload.get('owner_analytics', {}).get('owners', []):
            _row(o.get('name'), o.get('opportunities'),
                 o.get('pipeline_value'), o.get('won'), o.get('win_rate'))

        # --- Section 4: Productivity ------------------------------
        _section("Productivity")
        _row("User", "Leads Assigned", "Activities", "Conversion %", "Revenue")
        for u in payload.get('productivity', {}).get('users', []):
            _row(u.get('name'), u.get('leads_assigned'),
                 u.get('activities_completed'),
                 u.get('conversion_rate'), u.get('revenue_generated'))

        # --- Section 5: Activity Monthly --------------------------
        _section("Activity (This Month)")
        m = payload.get('activity', {}).get('monthly', {})
        _row("Type", "Count")
        for k in ('calls', 'meetings', 'emails', 'other',
                  'completed', 'scheduled'):
            _row(k.title(), m.get(k))

        # --- Section 6: Startup Metrics ---------------------------
        _section("Startup Health")
        s = payload.get('startup', {})
        bh = s.get('business_health', {})
        for k in ('last_booking_date', 'days_since_booking',
                  'days_since_won_opp', 'days_since_lead'):
            _row("Health / " + k.replace('_', ' ').title(), bh.get(k))
        g = s.get('growth', {})
        for k, label in [
            ('weekly_growth', 'Weekly Lead Growth %'),
            ('monthly_growth', 'Monthly Lead Growth %'),
            ('quarterly_growth', 'Quarterly Lead Growth %'),
            ('revenue_growth', 'Revenue Growth %'),
        ]:
            _row("Growth / " + label, g.get(k))
        r = s.get('risk', {})
        for k in ('no_bookings_7_days', 'no_bookings_14_days',
                  'no_bookings_30_days', 'overdue_followups', 'stagnant_opps'):
            _row("Risk / " + k.replace('_', ' ').title(), r.get(k))

        # --- Done --------------------------------------------------
        text = buf.getvalue()
        data = text.encode('utf-8')
        b64 = base64.b64encode(data)
        filename = "crm_dashboard_%s.csv" % _safe_filename(
            self.period or "all"
        )
        self.write({
            'file_data': b64,
            'filename': filename,
            'file_type': 'text/csv',
            'state': 'done',
        })
        return {'data': b64, 'filename': filename, 'file_type': 'text/csv'}

    # ------------------------------------------------------------------
    # XLSX — openpyxl
    # ------------------------------------------------------------------

    def _generate_xlsx(self):
        """Build a multi-sheet XLSX with one section per sheet."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise UserError(_("Excel export requires the openpyxl Python library."))

        payload = self._compute_payload()

        wb = openpyxl.Workbook()
        # Remove the default sheet
        wb.remove(wb.active)

        # Brand styles
        header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        header_fill = PatternFill(start_color="0B1F3A", end_color="0B1F3A", fill_type="solid")
        section_fill = PatternFill(start_color="D8C08C", end_color="D8C08C", fill_type="solid")
        section_font = Font(bold=True, color="0B1F3A", size=12, name="Calibri")
        cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

        def _header_row(ws, row, headers):
            for col_idx, h in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = cell_alignment

        def _section_title(ws, row, title):
            cell = ws.cell(row=row, column=1, value=title)
            cell.font = section_font
            cell.fill = section_fill

        def _auto_width(ws):
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for c in col:
                    try:
                        v = c.value
                        if v is None:
                            continue
                        s = str(v)
                        if len(s) > max_len:
                            max_len = len(s)
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        # --- Sheet 1: KPI Overview ----------------------------------
        ws = wb.create_sheet("KPI Overview")
        kpi = payload.get('kpi', {})
        _section_title(ws, 1, "Leads")
        _header_row(ws, 2, ["Metric", "Value"])
        leads = kpi.get('leads', {})
        for i, k in enumerate(('total', 'new_today', 'new_week', 'new_month',
                               'qualified', 'unqualified', 'awaiting_action'), start=3):
            ws.cell(row=i, column=1, value="Leads / " + k.replace('_', ' ').title())
            ws.cell(row=i, column=2, value=leads.get(k))
        row = 3 + len(('total', 'new_today', 'new_week', 'new_month',
                       'qualified', 'unqualified', 'awaiting_action')) + 1
        _section_title(ws, row, "Opportunities")
        _header_row(ws, row + 1, ["Metric", "Value"])
        opps = kpi.get('opportunities', {})
        for i, k in enumerate(('total', 'open', 'won', 'lost',
                               'created_today', 'created_week', 'created_month'),
                              start=row + 2):
            ws.cell(row=i, column=1, value="Opps / " + k.replace('_', ' ').title())
            ws.cell(row=i, column=2, value=opps.get(k))
        row = row + 2 + 7 + 1
        _section_title(ws, row, "Revenue")
        _header_row(ws, row + 1, ["Metric", "Value"])
        rev = kpi.get('revenue', {})
        for i, k in enumerate(('pipeline_value', 'forecast', 'won',
                               'month', 'quarter', 'year'), start=row + 2):
            ws.cell(row=i, column=1, value="Revenue / " + k.replace('_', ' ').title())
            ws.cell(row=i, column=2, value=rev.get(k))
        _auto_width(ws)

        # --- Sheet 2: Sales Funnel ---------------------------------
        ws = wb.create_sheet("Sales Funnel")
        _header_row(ws, 1, ["Stage", "Count", "Value", "Conversion %", "Drop-off %"])
        for i, s in enumerate(payload.get('funnel', {}).get('stages', []), start=2):
            ws.cell(row=i, column=1, value=s.get('name'))
            ws.cell(row=i, column=2, value=s.get('count'))
            ws.cell(row=i, column=3, value=s.get('value'))
            ws.cell(row=i, column=4, value=s.get('conversion'))
            ws.cell(row=i, column=5, value=s.get('drop_off'))
        _auto_width(ws)

        # --- Sheet 3: Pipeline by Owner -----------------------------
        ws = wb.create_sheet("Pipeline by Owner")
        _header_row(ws, 1, ["Owner", "Opportunities", "Pipeline Value", "Won", "Win Rate %"])
        for i, o in enumerate(payload.get('owner_analytics', {}).get('owners', []), start=2):
            ws.cell(row=i, column=1, value=o.get('name'))
            ws.cell(row=i, column=2, value=o.get('opportunities'))
            ws.cell(row=i, column=3, value=o.get('pipeline_value'))
            ws.cell(row=i, column=4, value=o.get('won'))
            ws.cell(row=i, column=5, value=o.get('win_rate'))
        _auto_width(ws)

        # --- Sheet 4: Productivity ---------------------------------
        ws = wb.create_sheet("Productivity")
        _header_row(ws, 1, ["User", "Leads", "Activities", "Conv %", "Revenue"])
        for i, u in enumerate(payload.get('productivity', {}).get('users', []), start=2):
            ws.cell(row=i, column=1, value=u.get('name'))
            ws.cell(row=i, column=2, value=u.get('leads_assigned'))
            ws.cell(row=i, column=3, value=u.get('activities_completed'))
            ws.cell(row=i, column=4, value=u.get('conversion_rate'))
            ws.cell(row=i, column=5, value=u.get('revenue_generated'))
        _auto_width(ws)

        # --- Sheet 5: Activity -------------------------------------
        ws = wb.create_sheet("Activity")
        _section_title(ws, 1, "This Month")
        _header_row(ws, 2, ["Type", "Count"])
        m = payload.get('activity', {}).get('monthly', {})
        for i, k in enumerate(('calls', 'meetings', 'emails', 'other',
                               'completed', 'scheduled'), start=3):
            ws.cell(row=i, column=1, value=k.title())
            ws.cell(row=i, column=2, value=m.get(k))
        _auto_width(ws)

        # --- Save to in-memory buffer -----------------------------
        buf = io.BytesIO()
        wb.save(buf)
        data = buf.getvalue()
        b64 = base64.b64encode(data)
        filename = "crm_dashboard_%s.xlsx" % _safe_filename(
            self.period or "all"
        )
        self.write({
            'file_data': b64,
            'filename': filename,
            'file_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'state': 'done',
        })
        return {
            'data': b64,
            'filename': filename,
            'file_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        }

    # ------------------------------------------------------------------
    # PDF — wkhtmltopdf against a QWeb template
    # ------------------------------------------------------------------

    def _generate_pdf(self):
        """Render the dashboard as a printable PDF using wkhtmltopdf.

        We render a QWeb HTML template (dashboard_pdf_template) into a
        temp file, then call wkhtmltopdf to convert it. The result is
        read back, base64-encoded, and stored on the wizard.
        """
        payload = self._compute_payload()
        # Render HTML via QWeb
        html = self.env['ir.qweb']._render(
            'crm_executive_dashboard.dashboard_pdf_template',
            values={
                'payload': payload,
                'period': self.period,
                'date_from': self.date_from,
                'date_to': self.date_to,
                'generated_at': fields.Datetime.now(),
                'company': self.env.company,
            },
        )

        # Find wkhtmltopdf (Odoo 19 renamed tools.find_executable -> tools.find_in_path)
        wkhtmltopdf_bin = tools.find_in_path('wkhtmltopdf') or '/usr/local/bin/wkhtmltopdf'
        if not wkhtmltopdf_bin:
            raise UserError(_(
                "wkhtmltopdf is not available on the server. "
                "Please install it or choose CSV/Excel format."
            ))

        # Write HTML to temp file, run wkhtmltopdf, read back
        with tempfile.NamedTemporaryFile(suffix='.html', delete=False) as in_f:
            if isinstance(html, bytes):
                in_f.write(html)
            else:
                in_f.write(html.encode('utf-8'))
            in_path = in_f.name
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as out_f:
            out_path = out_f.name

        try:
            cmd = [
                wkhtmltopdf_bin,
                '--quiet',
                '--enable-local-file-access',
                '--print-media-type',
                '--no-outline',
                '--margin-top', '10mm',
                '--margin-bottom', '10mm',
                '--margin-left', '10mm',
                '--margin-right', '10mm',
                '--page-size', 'A4',
                '--orientation', 'Landscape',
                in_path,
                out_path,
            ]
            proc = subprocess.run(
                cmd,
                capture_output=True,
                timeout=60,
            )
            if proc.returncode != 0:
                _logger.error("wkhtmltopdf failed: %s", proc.stderr.decode('utf-8', 'replace'))
                raise UserError(_("PDF generation failed (wkhtmltopdf exit %s). Try CSV/Excel instead.") % proc.returncode)
            with open(out_path, 'rb') as f:
                data = f.read()
        finally:
            try: import os; os.unlink(in_path)
            except Exception: pass
            try: import os; os.unlink(out_path)
            except Exception: pass

        b64 = base64.b64encode(data)
        filename = "crm_dashboard_%s.pdf" % _safe_filename(
            self.period or "all"
        )
        self.write({
            'file_data': b64,
            'filename': filename,
            'file_type': 'application/pdf',
            'state': 'done',
        })
        return {'data': b64, 'filename': filename, 'file_type': 'application/pdf'}

    # ------------------------------------------------------------------
    # Helper: compute the dashboard payload (delegates to KPI engine)
    # ------------------------------------------------------------------

    def _compute_payload(self):
        """Call the KPI engine with the wizard's filter and return raw data.

        We use the model's collect_all() (added in batch 1) to avoid
        duplicating business logic. The transient model is created in
        sudo mode because the user has already been authorised.
        """
        filter_dict = {}
        if self.period:
            filter_dict['period'] = self.period
        if self.date_from:
            filter_dict['date_from'] = fields.Date.to_string(self.date_from)
        if self.date_to:
            filter_dict['date_to'] = fields.Date.to_string(self.date_to)
        for k in ('user_id', 'team_id', 'source_id', 'stage_id', 'company_id'):
            v = getattr(self, k)
            if v:
                filter_dict[k] = v.id
        # Use sudo() so the engine sees all data, but only if the user
        # has the manager group. Regular users still get a meaningful
        # export (filtered to what they can see).
        # Odoo 19: sudo() lives on BaseModel, not Environment.
        kpi_model = self.env['crm.dashboard.kpi']
        if self.env.user.has_group('crm_executive_dashboard.group_crm_dashboard_manager'):
            kpi_model = kpi_model.sudo()
        return kpi_model.collect_all(filter_dict)


def _to_csv_cell(v):
    """Render a value for CSV output."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, datetime):
        return v.isoformat()
    return v
